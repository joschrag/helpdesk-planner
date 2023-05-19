"""Script to fill excel sheet according to data provided in other excel sheet."""
import pandas as pd
import xlwings as xw
import pathlib
import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
import datetime as dt
from typing import Any
import time
class Termin:
    """Class containing information about the events."""

    instances = set()

    def __init__(
        self,
        wt: str,
        start: dt.time,
        finish: dt.time,
        tutor: str,
        sp: tuple[str,str],
        ort: str,
    ) -> None:
        """Initialize the Termin class.

        Args:
            wt (str): weekday
            start (dt.time): start time
            finish (dt.time): end time
            tutor (str): name of tutor
            sp (str): knowledge area of tutor
            ort (str): location
        """
        self.wt = wt
        self.tutor = tutor
        self.start = start
        self.finish = finish
        self.sp = sp
        self.ort = ort
        self.deu = []
        self.__class__.instances.add(self)
        self.cols = self.day_to_cols()
        self.rows = self.time_to_row()
        self.hasmoved = False
        self.isLeft = False


    def __str__(self) -> str:
        """Get string representation.

        Returns:
            str: string description.
        """
        return "\n"+"\n".join([f"{pair[0]}: {pair[1]}" for pair in list(vars(self).items())])

    def check_overlap(self) -> None:
        """Check if other Class object overlap timewise."""
        for termin in self.instances:
            if not self == termin and termin.wt == self.wt:
                if (
                    (termin.start <= self.start < termin.finish)
                    or (termin.start < self.finish <= termin.finish)
                    or (self.start <= termin.start < self.finish)
                    or (self.start < termin.finish <= self.finish)
                ):
                    self.deu.append(termin)


    def correct_cols(self) -> None:
        """Correct columns if overlap exists."""
        for termin in self.deu:
            if termin.hasmoved:
                self.isLeft = not termin.isLeft
                self.hasmoved = True
            else:
                self.isLeft = True
                self.hasmoved = True

        if self.isLeft:
            self.cols = (self.cols[0], self.cols[0])
        else:
            self.cols = (self.cols[1], self.cols[1])

    def day_to_cols(self) -> tuple:
        """Translate weekdays into column numbers.

        Returns:
            tuple: (start_col, end_col)
        """
        m_dict = {
            "Montag": (4, 5),
            "Dienstag": (6, 7),
            "Mittwoch": (8, 9),
            "Donnerstag": (10, 11),
            "Freitag": (12, 13),
        }
        return m_dict[self.wt]

    def time_to_row(self) -> tuple:
        """Translate start and finish times into rows.

        Returns:
            tuple: (start_row, end_row)
        """
        finish_dt = dt.datetime.combine(dt.datetime.now().date(), self.finish)
        start_dt = dt.datetime.combine(dt.datetime.now().date(), self.start)
        first_row = dt.datetime.combine(dt.datetime.now().date(), dt.time(8, 0))
        lower_diff = start_dt - first_row
        upper_diff = finish_dt - first_row

        start_row = (
            lower_diff.seconds // 3600 * 2
            + (lower_diff.seconds % 3600) // 1800
            + 5
        )
        end_row = (
            upper_diff.seconds // 3600 * 2
            + (upper_diff.seconds % 3600) // 1800
            + 4
        )

        return (start_row, end_row)

    def fill_colors(self, ws: Worksheet) -> None:
        """Fill the excel sheet with the corresponding color.

        Args:
            ws (Worksheet): output Worksheet
        """
        cols = {"Rüsselsheim": "E2EFDA", "WBS": "DDEBF7", "online": "FFF2CC"}
        color = cols[self.ort]
        for rows in ws.iter_rows(
            min_row=self.rows[0],
            max_row=self.rows[1],
            min_col=self.cols[0],
            max_col=self.cols[1],
        ):
            for cell in rows:
                cell.fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    def add_border(self, ws: Worksheet) -> None:
        """Add borders into excel worksheet.

        Args:
            ws (Worksheet): output Worksheet
        """
        med = Side(border_style="medium", color="000000")
        dot = Side(border_style="hair", color="000000", style="dotted")
        for i in range(self.rows[0], self.rows[1] + 1):
            for j in range(self.cols[0], self.cols[1] + 1):
                cell = ws.cell(i, j)
                if i == self.rows[0]:
                    if j == self.cols[0]:
                        if j == self.cols[1]:
                            cell.border = Border(top=med, left=med, right=med)
                        else:
                            cell.border = Border(top=med, left=med)
                    elif j == self.cols[1]:
                        cell.border = Border(top=med, right=med)
                if i == self.rows[1]:
                    if j == self.cols[0]:
                        if j == self.cols[1]:
                            cell.border = Border(
                                bottom=med, left=med, right=med
                            )
                        else:
                            cell.border = Border(bottom=med, left=med)
                    elif j == self.cols[1]:
                        cell.border = Border(bottom=med, right=med)
                if not i == self.rows[0] and not i == self.rows[1]:
                    if j == self.cols[0]:
                        if j == self.cols[1]:
                            cell.border = Border(
                                left=med, right=med, bottom=dot, top=dot
                            )
                        else:
                            cell.border = Border(left=med, bottom=dot, top=dot)
                    elif j == self.cols[1]:
                        cell.border = Border(right=med, bottom=dot, top=dot)

    def add_desc(self, ws: Worksheet):
        """Add description text to the event in the output worksheet.

        Args:
            ws (Worksheet): output worksheet
        """
        room_dict = {"Rüsselsheim": "G007", "WBS": "II-02", "online": "online"}
        name_cell = ws.cell(self.rows[0], self.cols[0])
        exp_cell = ws.cell(self.rows[0] + 1, self.cols[0])
        exp2_cell = ws.cell(self.rows[1] + 2, self.cols[0])
        ort_cell = ws.cell(self.rows[1], self.cols[0])
        name_cell.value = self.tutor
        name_cell.font = Font(bold=True)
        if self.rows[1] - self.rows[0] > 2:
            exp_cell.value, exp2_cell.value = self.sp
        else:
            exp_cell.value = "".join(self.sp)
        exp_cell.font = Font(bold=False)
        exp2_cell.font = Font(bold=False)
        ort_cell.value = room_dict[self.ort]
        ort_cell.font = Font(bold=True)



def load_template() -> pathlib.Path:
    """Load the template excel worksheet.

    Returns:
        pathlib.Path: Path to the ouput file.
    """
    path1 = pathlib.Path(__file__).parents[1] / "Template.xlsx"
    path2 = pathlib.Path(__file__).parents[1] / "Ergebnis.xlsx"

    wb1 = xw.Book(path1)
    wb2 = xw.Book()

    ws1 = wb1.sheets(1)
    ws1.copy(before=wb2.sheets[0])
    wb2.save(path2)
    wb2.app.quit()

    wb = xl.load_workbook(path2)
    sheet_names = wb.sheetnames
    for sheet in sheet_names[1:]:
        wb.remove(wb[sheet])
    ws = wb[sheet_names[0]]
    ws.title = "Wochenplan"
    wb.save(path2)
    return path2

def write_text(inpath:pathlib.Path,ws:Worksheet) -> None:
    """Write the timeframe of the plan into the excel.

    Args:
        inpath (pathlib.Path): input path for the date
        ws (Worksheet): output worksheet
    """    
    df = pd.read_excel(inpath, sheet_name="KW")
    date:dt.datetime = pd.to_datetime(df.iloc[0,0])# type: ignore    
    cal_week_str = date.strftime("%V")
    cal_week_date = date.strftime("%G %V")
    startdate = time.asctime(time.strptime(f"{cal_week_date} 0", "%Y %W %w")) 
    startdate = dt.datetime.strptime(startdate, "%a %b %d %H:%M:%S %Y")
    friday = startdate - dt.timedelta(days=2)
    monday = startdate - dt.timedelta(days=6)
    friday_str = friday.strftime("%d.%m.%Y")
    monday_str = monday.strftime("%d.%m.%Y")
    out_str = f"Helpdeskplan für KW {cal_week_str} von {monday_str} bis {friday_str}."
    ws.cell(2,1).value = out_str


if __name__ == "__main__":
    print("Reading input data -> 'Liste.xlsm'")
    inpath = pathlib.Path(__file__).parents[1] / "Liste.xlsm"
    df = pd.read_excel(inpath, sheet_name="Liste")
    df = df.dropna(subset="Anfangszeit")
    print("Creating classes.")
    for index, row in df.iterrows():
        
        day, start, end, tutor, sp1, sp2, location = row
        sp : tuple = (sp1,sp2)
        Termin(day,start,end,tutor,sp,location)
    for t in Termin.instances:
        t.check_overlap()
    print("Loading template -> 'Template.xlsx'")
    path = load_template()
    wb = xl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    for t in Termin.instances:
        if len(t.deu) > 0:
            t.correct_cols()
    for t in Termin.instances:
        t.fill_colors(ws)
        t.add_border(ws)
        t.add_desc(ws)
    print("Writing data to file -> Ergebnis.xlsx")
    write_text(inpath,ws)

    wb.save(path)
    print("Created plan. Output in file -> Ergebnis.xlsx")
